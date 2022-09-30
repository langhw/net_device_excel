# -*- coding: UTF-8 -*-
'''
@Project ：net_device_excel 
@File ：H3CFw_Rules_Biao.py
@Author ：lianghongwei
@Date ：2022/9/28 10:42 
@PRODUCT : PyCharm
'''

import os
from datetime import datetime
from multiprocessing.pool import ThreadPool
import openpyxl
import threading
from netmiko import ConnectHandler as ch
from netmiko.ssh_exception import (NetMikoTimeoutException, AuthenticationException, SSHException)
# from prettytable import PrettyTable
import re
# import chardet
import pandas as pd
from pandas import DataFrame as df
from openpyxl import styles
# from styleframe import StyleFrame, Styler

column = ['ID', 'Source-Zone', 'Destination-Zone', 'S_ip_name', 'S_ip', 'D_ip_name', 'D_ip',
                       'Service_name', 'Protocol', 'Port', 'Description', 'Action', 'State']
dir = r'C:\Users\lianghongwei-hzgs\Desktop\12'
log = 'JG'
if not os.path.exists(log): os.mkdir(log)
logtime = datetime.now().strftime('%Y-%m-%d_%H')
dirpath = os.path.join(log, logtime + '_FWrules.xlsx')
pool = ThreadPool(5)  # 并发数
queueLock = threading.Lock()  # 线程锁
logtime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  # 时间
success = []
fail = []
cmds = ['display current-configuration']
dev_info_xlsx = 'device_info.xlsx'

def device_info(xlsx=dev_info_xlsx):
    """在表格中获取设备登录信息"""
    book = openpyxl.load_workbook(xlsx)
    sheet = book.active
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=6):
        if row[0].value is not None:
            continue
        ip = row[2].value
        username = 'admin' if row[3].value is None else row[3].value
        password = 'admin' if row[4].value is None else row[4].value
        info_dict = {'ip': ip,
                     # 'protocol': 'ssh',
                     # 'port': '22',
                     'username': username,
                     'password': password,
                     'device_type': 'hp_comware',
                     }
        yield info_dict
    book.close()

def denglu(dev_info):
    """ssh登录设备"""
    global fail
    try:
        conn = ch(**dev_info)
        return conn
    except NetMikoTimeoutException:
        e = "Failed.....{:<15} 连通性问题!".format(dev_info['ip'])
        printPretty(e)
        fail.append(dev_info['ip'])

    except AuthenticationException:
        e = "Failed.....{:<15} 用户名或密码错误!".format(dev_info['ip'])
        printPretty(e)
        fail.append(dev_info['ip'])

    except SSHException:
        e = "Failed.....{:<15} SSH版本不兼容!".format(dev_info['ip'])
        printPretty(e)
        fail.append(dev_info['ip'])

    except Exception as e:
        e = "Failed.....{:<15} connectHandler Error: {}".format(dev_info['ip'], e)
        printPretty(e)
        fail.append(dev_info['ip'])

def get_conf(**device_inf):
    """获取防火墙配置信息"""
    global success
    printPretty('设备...{:.<15}...开始执行'.format(device_inf['ip']))
    conn = denglu(device_inf)
    output = ''
    if conn:
        # 获取设备名称并格式化
        # hostname = format_hostname(conn.find_prompt(), dev_info['device_type'])
        hostname = format_hostname(conn.find_prompt())
        try:
            for cmd in cmds:
                output += conn.send_command(cmd, strip_prompt=False, strip_command=False, )
            success.append(device_inf['ip'])

        except Exception as e:
            output = f"run Failed...{device_inf['ip']} : {e}"
            printPretty(output)
            fail.append(device_inf['ip'])
        finally:
            # 退出netmiko session
            conn.disconnect()
        return hostname, output

def printPretty(msg):
    """打印消息"""
    # 在并发的场景中，避免在一行打印出多个结果，不方便查看
    queueLock.acquire()  # 加锁
    print(msg)
    queueLock.release()  # 释放锁

def printSum(msg):
    """打印结果汇总信息"""
    global success, fail
    total_devices, success, fail = len(success + fail), len(success), len(fail)
    total_time = "{:0.2f}s".format(msg.total_seconds())
    # tb = PrettyTable(['设备总数', '成功', '失败', '总耗时'])
    # tb.add_row([total_devices, success, fail, total_time])
    print('设备总数:{}, 成功:{}, 失败:{}, 总耗时:{}'.format(total_devices, success, fail, total_time))
    # print(tb)

def format_hostname(hostname):
    """格式化主机名称"""
    new_hostname = hostname.split()[0].strip("<>#$() ")
    return new_hostname

def log_dir():
    """创建目录"""
    # 判断当前目录是否有LOG文件夹，不存在则创建
    if not os.path.exists('LOG'):
        os.makedirs('LOG')
    return 'LOG'

# def get_info():
#     """读取配置文件，并将文件格式化到列表中"""
#     for file_n in get_conf():
#         file_list = file_n[1].split('\n#\n')
#         file_name = file_n[0]
#         yield file_name, file_list

def pp_data(file):
    """匹配关键信息到字典中"""
    # 定义匹配地址对象
    pd_object_a_n = re.compile(r'object-group (?:ip|ipv6) address (.*)')
    pd_object_a = re.compile(r'(?: \d+ network (?:host address|subnet|host name|range) (.*))+')
    # 定义匹配服务对象
    pd_object_s_n = re.compile(r'object-group service (.*)')
    # pd_object_s = re.compile(r'(?: \d+ service (\w+) destination (\w+) (\d+))+')
    pd_object_s = re.compile(r'(?: \d+ service (\w+) destination (.*))+')
    # 定义匹配安全策略
    pd_rule_name = re.compile(r'(\d+) name (.*)')
    pd_rule_desc = re.compile(r'  description (.*)')
    pd_rule_action = re.compile(r'  action (\w+)')
    pd_rule_state = re.compile(r'  (disable)')
    pd_rule_sz = re.compile(r'  source-zone (\w+)')
    pd_rule_dz = re.compile(r'  destination-zone (\w+)')
    pd_rule_si = re.compile(r'  source-ip (.*)')
    pd_rule_di = re.compile(r'  destination-ip (.*)')
    pd_rule_s = re.compile(r'  service (.*)')
    #
    file_name = file[0]
    file_log = file[1].split('\n#\n')
    ob_a_t = {'any': 'any'}  # 对象地址字典
    ob_s_t = {'any': ['any', 'any']}  # 对象服务字典
    rule_t = {}  # 安全策略字典
    for line in file_log:
        if 'object-group ip address' in line or 'object-group ipv6 address' in line:
            ob_a_n = pd_object_a_n.findall(line)[0]  # 找到地址对象的名称
            ob_a = pd_object_a.findall(line)  # 找到地址对象的地址信息
            n = 0
            if len(ob_a) != 0:
                while n < len(ob_a):
                    if n == 0:
                        ob_a_t[ob_a_n] = ob_a[n]
                    else:
                        ob_a_t[ob_a_n] = ob_a_t[ob_a_n] + chr(10) + ob_a[n]
                    n += 1
            else:
                ob_a_t[ob_a_n] = 'None'
        elif 'object-group service' in line:
            ob_s_n = pd_object_s_n.findall(line)[0]  # 找到服务对象的名称
            ob_s = pd_object_s.findall(line)  # 找到服务对象的服务信息，包括协议类型，协议指定，协议端口号
            ob_s_t[ob_s_n] = []  # 服务对象为列表，0是协议，1是端口
            n = 0
            if len(ob_s) != 0:
                while n < len(ob_s):
                    if n == 0:
                        ob_s_t[ob_s_n].append(ob_s[n][0])
                        ob_s_t[ob_s_n].append(ob_s[n][1])
                    else:
                        ob_s_t[ob_s_n][0] = ob_s_t[ob_s_n][0] + chr(10) + ob_s[n][0]
                        ob_s_t[ob_s_n][1] = ob_s_t[ob_s_n][1] + chr(10) + ob_s[n][1]
                    n += 1
            else:
                ob_s_t[ob_s_n].append('None')
                ob_s_t[ob_s_n].append('None')
        elif 'security-policy ip' in line:
            rule_line = line.split(' rule ')
            for rule in rule_line[1:]:
                n1 = n2 = n3 = n4 = n5 = 0
                rule_id = pd_rule_name.findall(rule)[0][0]
                rule_name = pd_rule_name.findall(rule)[0][1]
                rule_desc = pd_rule_desc.findall(rule)
                rule_action = pd_rule_action.findall(rule)
                rule_state = pd_rule_state.findall(rule)
                rule_sz = pd_rule_sz.findall(rule)
                rule_dz = pd_rule_dz.findall(rule)
                rule_si = pd_rule_si.findall(rule)
                rule_di = pd_rule_di.findall(rule)
                rule_s = pd_rule_s.findall(rule)
                #
                if len(rule_desc) != 0:
                    rule_desc = rule_desc[0]
                else:
                    rule_desc = ''
                if len(rule_action) != 0:
                    rule_action = rule_action[0]
                else:
                    rule_action = 'Drop'
                if len(rule_state) != 0:
                    rule_state = rule_state[0]
                else:
                    rule_state = 'enable'
                if len(rule_sz) == 0:
                    rule_sz = ['any']
                if len(rule_dz) == 0:
                    rule_dz = ['any']
                if len(rule_si) == 0:
                    rule_si = ['any']
                if len(rule_di) == 0:
                    rule_di = ['any']
                if len(rule_s) == 0:
                    rule_s = ['any']
                #
                while n1 < len(rule_sz):
                    if n1 == 0:
                        rule_sz_m = rule_sz[n1]
                    else:
                        rule_sz_m = rule_sz_m + chr(10) + rule_sz[n1]
                    n1 += 1
                while n2 < len(rule_dz):
                    if n2 == 0:
                        rule_dz_m = rule_dz[n2]
                    else:
                        rule_dz_m = rule_dz_m + chr(10) + rule_dz[n2]
                    n2 += 1
                while n3 < len(rule_si):
                    if n3 == 0:
                        rule_si_m = [rule_si[n3], ob_a_t[rule_si[n3]]]
                    else:
                        rule_si_m[0] = rule_si_m[0] + chr(10) + '-' * 10 + chr(10) + rule_si[n3]
                        rule_si_m[1] = rule_si_m[1] + chr(10) + '-' * 10 + chr(10) + ob_a_t[rule_si[n3]]
                    n3 += 1
                while n4 < len(rule_di):
                    if n4 == 0:
                        rule_di_m = [rule_di[n4], ob_a_t[rule_di[n4]]]

                    else:
                        rule_di_m[0] = rule_di_m[0] + chr(10) + '-' * 10 + chr(10) + rule_di[n4]
                        rule_di_m[1] = rule_di_m[1] + chr(10) + '-' * 10 + chr(10) + ob_a_t[rule_di[n4]]
                    n4 += 1
                while n5 < len(rule_s):
                    if rule_s[n5] not in ob_s_t:
                        ob_s_t[rule_s[n5]] = ['Predefined', rule_s[n5]]
                    if n5 == 0:
                        rule_s_m = [rule_s[n5], ob_s_t[rule_s[n5]][0], ob_s_t[rule_s[n5]][1]]
                    else:
                        rule_s_m[0] = rule_s_m[0] + chr(10) + '-' * 10 + chr(10) + rule_s[n5]
                        rule_s_m[1] = rule_s_m[1] + chr(10) + '-' * 10 + chr(10) + ob_s_t[rule_s[n5]][0]
                        rule_s_m[2] = rule_s_m[2] + chr(10) + '-' * 10 + chr(10) + ob_s_t[rule_s[n5]][1]
                    n5 += 1
                #
                rule_t[rule_name] = []
                rule_t[rule_name].append(rule_id)
                rule_t[rule_name].append(rule_sz_m)
                rule_t[rule_name].append(rule_dz_m)
                rule_t[rule_name].append(rule_si_m[0])
                rule_t[rule_name].append(rule_si_m[1])
                rule_t[rule_name].append(rule_di_m[0])
                rule_t[rule_name].append(rule_di_m[1])
                rule_t[rule_name].append(rule_s_m[0])
                rule_t[rule_name].append(rule_s_m[1])
                rule_t[rule_name].append(rule_s_m[2])
                rule_t[rule_name].append(rule_desc)
                rule_t[rule_name].append(rule_action)
                rule_t[rule_name].append(rule_state)
    return file_name, rule_t

def write_excel(pipei):
    """主程序，将信息写入表格保持"""
    start_time = datetime.now()
    # writer = StyleFrame.ExcelWriter(dirpath)
    writer = pd.ExcelWriter(dirpath)
    for pp in pipei:
        data = df.from_dict(pp[1], orient='index', columns=column)
        data.reset_index(inplace=True)
        # data.index = data.index + 1
        data.rename(columns={'index': 'Rule_name'}, inplace=True)
        # column1 = ['Rule_name'] + column
        # sf = StyleFrame(data)
        # sf.apply_column_style(cols_to_style=column,
        #                       styler_obj=Styler(horizontal_alignment='left'),
        #                       style_header=False)
        # sf.to_excel(
        #     excel_writer=writer,
        #     sheet_name=pp[0],
        #     best_fit=column1,
        #     columns_and_rows_to_freeze='B2',
        #     row_to_add_filters=0,
        # )
        data.to_excel(excel_writer=writer, sheet_name=pp[0], index=False,)
        writer.save()
    writer.close()
    end_time = datetime.now()
    print('-' * 50)
    # print('>>>>所有已经执行完成，总共耗时{:0.2f}秒.<<<'.format((end_time - start_time).total_seconds()))
    printSum(end_time - start_time)

def meihua(xlsx):
    font = styles.Font(name='微软雅黑', size=11, )

    border = styles.Border(
        left=styles.Side(border_style='thin', color='FF000000'),
        right=styles.Side(border_style='thin', color='FF000000'),
        top=styles.Side(border_style='thin', color='FF000000'),
        bottom=styles.Side(border_style='thin', color='FF000000')
    )

    alignment = styles.Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True,
    )

    wl = openpyxl.load_workbook(xlsx)
    wb_list = wl.sheetnames
    print(wb_list)
    a = wl.active.max_row
    b = wl.active.max_column
    print(a, b)
    for wb in wb_list:
        wa = wl[wb]
        print(wa.title)
        wa.delete_cols(1)
        mr = wa.max_row
        mrf = 'A1:' + 'N' + str(mr)
        wa.freeze_panes = 'C2'
        wa.column_dimensions
        for key in list(wa._cells.keys()):
            wa._cells[key].alignment = alignment
            wa._cells[key].font = font
            wa._cells[key].border = border
        wa.row_dimensions
        wa.column_dimensions
        wa.auto_filter.ref = mrf
    wl.save(xlsx)
    wl.close()


def main():
    writer = pd.ExcelWriter(dirpath)
    for dev_info in device_info():
        getconf = get_conf(**dev_info)
        pipei = pp_data(getconf)
        data = df.from_dict(pipei[1], orient='index', columns=column)
        data.reset_index(inplace=True)
        data.index = data.index + 1
        data.rename(columns={'index': 'Rule_name'}, inplace=True)
        data.to_excel(excel_writer=writer, sheet_name=pipei[0])
        writer.save()
    writer.close()
    meihua(dirpath)
if __name__ == "__main__":
    start_time = datetime.now()
    main()
    end_time = datetime.now()
    print('>>>>所有已经执行完成，总共耗时{:0.2f}秒.<<<'.format((end_time - start_time).total_seconds()))
